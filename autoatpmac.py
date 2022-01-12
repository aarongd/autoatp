#for mac
import openpyxl
from appscript import app, k
from mactypes import Alias
from pathlib import Path
import tkinter as tk
from tkinter import *


#Where information pulled from the Action Tracker, SLA Tracker, and cost org will go.
li = {
    "Candidate Name" : "",
    "Company" : "",
    "JITR" : "",
    "CSR" : "",
    "Labor Category" : "",
    "Level" : "",
    "Effective Date" : "",
    "Submitted Rate to CACI" : "",
    "SLA" : "",
    "CLIN" : "0820",
    "Resource ID" : "",
    "Cost Center" : "",
}


def show_entry_fields():
    print("CSR Number: %s\nResource ID: %s" % (e1.get(), e2.get()))
    e1.delete(0, tk.END)
    e2.delete(0, tk.END)

input_values = []

def send_entry_fields():
        input_values.append(e1.get())
        input_values.append(e2.get())
        master.quit()

master = tk.Tk()
master.title('ATP Notification')
tk.Label(master, text="CSR Number").grid(row=0)
tk.Label(master, text="Resource ID").grid(row=1)

e1 = tk.Entry(master)
e2 = tk.Entry(master)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)
e1.insert(10, "####-#####")
e1.configure(state=DISABLED)
e2.insert(10, "#########")
e2.configure(state=DISABLED)

def on_click(event):
    e1.configure(state=NORMAL)
    e2.configure(state=NORMAL)
    e1.delete(0, END)
    e2.delete(0, END)

    # make the callback only work once
    e1.unbind('<Button-1>', on_click_id)
    e2.unbind('<Button-1>', on_click_id2)

on_click_id = e1.bind('<Button-1>', on_click)
on_click_id2 = e2.bind('<Button-1>', on_click)

tk.Button(master,
          text='Quit', command=master.quit).grid(row=3,
                                    column=0,
                                    sticky=tk.W,
                                    pady=4)


tk.Button(master, text='Send', command=send_entry_fields).grid(row=3,
                                                               column=1,
                                                               sticky=tk.W,
                                                               pady=4)

master.mainloop()
tk.mainloop()

#the two inputs needed to put in manually
#csr_input = input("CSR Number: ")
#csr_input = 2019-33333
csr_input = input_values[0]
#resource_id_input = input("Resource ID: ")
resource_id_input = input_values[1]
li["Resource ID"] = resource_id_input

#opens the Action Tracker workbook
path = "/Users/aaron/Desktop/action tracker.xlsx"
wb = openpyxl.load_workbook(path)
sheets = wb.sheetnames
ws = wb.active
#n=0
#ws = wb[sheets[n]].active

#to find all rows within the CSR column in the Action tab
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        if cell.value == csr_input:
            #saving all relevant values from this excel book to memory
            full_name = ws.cell(row=cell.row, column=6).value
            company = ws.cell(row=cell.row, column=9).value
            jitr = ws.cell(row=cell.row, column=2).value
            labor_category = ws.cell(row=cell.row, column=7).value
            level = ws.cell(row=cell.row, column=8).value
            start_date = ws.cell(row=cell.row, column=18).value
            start_date2 = start_date.strftime("%d/%m/%Y")
            rate2caci = ws.cell(row=cell.row, column=10).value
            #start adding all saved values from this excel book to the list
            li["Candidate Name"] = full_name
            li["Company"] = company
            li["JITR"] = jitr
            li["CSR"] = csr_input
            li["Labor Category"] = labor_category
            li["Level"] = level
            li["Effective Date"] = start_date2
            li["Submitted Rate to CACI"] = rate2caci
            # Cost center is a constant unless for a few specific JITRs
            if jitr in (1124, 1125, 1126, 1158, 1160, 1166):
                li["Cost Center"] = 3373
            else:
                li["Cost Center"] = 3393
        else:
            print("CSR value not found")

#opens the SLA tracker
path2 = "/Users/aaron/Desktop/sla tracker.xlsx"

wb2 = openpyxl.load_workbook(path2)
sheets2 = wb2.sheetnames
ws2 =wb2.active
#n=0
#ws = wb[sheets[n]].active

for row in ws2.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        jitr_value = li['JITR']
        if cell.value == jitr_value:
            #saving all relevant values from this excel book to memory
            sla = ws2.cell(row=cell.row, column=2).value
            #start adding all saved values from this excel book to the list
            li["SLA"] = sla

print(li)

#to send the email
class Outlook(object):
    def __init__(self):
        self.client = app('Microsoft Outlook')

class Message(object):
    def __init__(self, parent=None, subject='', body='', to_recip=['aarongd1995@gamil.com'], cc_recip=[''], show_=True):
        if parent is None: parent = Outlook()
        client = parent.client

        self.msg = client.make(
            new=k.outgoing_message,
            with_properties={k.subject: subject, k.content: body}
        )

        self.add_recipients(emails=to_recip, type_='to')
        self.add_recipients(emails=cc_recip, type_='cc')

        if show_: self.show()

    def show(self):
        self.msg.open()
        self.msg.activate()

    def add_recipients(self, emails, type_='to'):
        if not isinstance(emails, list): emails = [emails]
        for email in emails:
            self.add_recipient(email=email, type_=type_)

    def add_recipient(self, email, type_='to'):
        msg = self.msg

        if type_ == 'to':
            recipient = k.to_recipient
        elif type_ == 'cc':
            recipient = k.cc_recipient

        msg.make(new=recipient, with_properties={k.email_address: {k.address: email}})


