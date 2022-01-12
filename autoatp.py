import openpyxl
import win32com.client as win32
import tkinter as tk
from tkinter import *

#Where information pulled from the Action Tracker, SLA Tracker, and cost org will go.
li = {
    "Candidate Name" : "",
    "Company" : "",
    "JITR" : "",
    "CSR" : "",
    "Labor Category" : "",
    "Level" : "",
    "Effective Date" : "",
    "Submitted Rate to CACI" : "",
    "SLA" : "",
    "CLIN" : "0820",
    "Resource ID" : "",
    "Cost Center" : "",
}

def show_entry_fields():
    print("CSR Number: %s\nResource ID: %s" % (e1.get(), e2.get()))
    e1.delete(0, tk.END)
    e2.delete(0, tk.END)

input_values = []

def send_entry_fields():
        input_values.append(e1.get())
        input_values.append(e2.get())
        master.quit()

master = tk.Tk()
master.title('ATP Notification')
tk.Label(master, text="CSR Number").grid(row=0)
tk.Label(master, text="Resource ID").grid(row=1)

e1 = tk.Entry(master)
e2 = tk.Entry(master)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)
e1.insert(10, "####-######")
e1.configure(state=DISABLED)
e2.insert(10, "#########")
e2.configure(state=DISABLED)

def on_click(event):
    e1.configure(state=NORMAL)
    e2.configure(state=NORMAL)
    e1.delete(0, END)
    e2.delete(0, END)

    # make the callback only work once
    e1.unbind('<Button-1>', on_click_id)
    e2.unbind('<Button-1>', on_click_id2)

on_click_id = e1.bind('<Button-1>', on_click)
on_click_id2 = e2.bind('<Button-1>', on_click)

tk.Button(master,
          text='Quit',
          command=master.quit).grid(row=3,
                                    column=0,
                                    sticky=tk.W,
                                    pady=4)


tk.Button(master, text='Send', command=send_entry_fields).grid(row=3,
                                                               column=1,
                                                               sticky=tk.W,
                                                               pady=4)

master.mainloop()
tk.mainloop()

#the two inputs needed to put in manually
#csr_input = input("CSR Number: ")
#csr_input = 2019-33333
csr_input = input_values[0]
#resource_id_input = input("Resource ID: ")
resource_id_input = input_values[1]
li["Resource ID"] = resource_id_input

#opens the Action Tracker workbook
path = "C:\\Users\\Aaron\\Desktop\\action tracker.xlsx"

wb = openpyxl.load_workbook(path)
sheets = wb.sheetnames
ws = wb.active
#n=0
#ws = wb[sheets[n]].active

#to find all rows within the CSR column in the Action tab
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        if cell.value == csr_input:
            #saving all relevant values from this excel book to memory
            full_name = ws.cell(row=cell.row, column=6).value
            company = ws.cell(row=cell.row, column=9).value
            jitr = ws.cell(row=cell.row, column=2).value
            labor_category = ws.cell(row=cell.row, column=7).value
            level = ws.cell(row=cell.row, column=8).value
            start_date = ws.cell(row=cell.row, column=18).value
            rate2caci = ws.cell(row=cell.row, column=10).value
            #start adding all saved values from this excel book to the list
            li["Candidate Name"] = full_name
            li["Company"] = company
            li["JITR"] = jitr
            li["CSR"] = csr_input
            li["Labor Category"] = labor_category
            li["Level"] = level
            li["Effective Date"] = start_date
            li["Submitted Rate to CACI"] = rate2caci
            # Cost center is a constant unless for a few specific JITRs
            if jitr in (1124, 1125, 1126, 1158, 1160, 1166):
                li["Cost Center"] = 3373
            else:
                li["Cost Center"] = 3393
        else:
            print("CSR value not found")

#opens the SLA tracker
#if windows
path2 = "C:\\Users\\Aaron\\Desktop\\sla tracker.xlsx"

wb2 = openpyxl.load_workbook(path2)
sheets2 = wb2.sheetnames
ws2 =wb2.active
#n=0
#ws = wb[sheets[n]].active

for row in ws2.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        jitr_value = li['JITR']
        if cell.value == int(jitr_value):
            #saving all relevant values from this excel book to memory
            sla = ws2.cell(row=cell.row, column=2).value
            #start adding all saved values from this excel book to the list
            li["SLA"] = sla


print(li)

#to send the email
outlook = win32.Dispatch('outlook.application')
mail = outlook.CreateItem(0)
mail.To = 'Aarongd1995@gmail.com'
mail.CC = 'Aarongd1995@gmail.com'
mail.Subject = 'ATP Notification - ' + li['Candidate Name']
print(mail.Subject)
mail.Body = 'Kim,\n\n' \
            'Plase find the ATP Notification below:\n\n'\
            'Candidate Name: ' + li['Candidate Name']+ '\n'\
            'Postion #: JITR ' + li['JITR'] + ' / ' + li['CSR']+'\n'\
            'Labor Category: ' + li['Labor Category'] + '\n'\
            'Level: ' + li['Level'] + '\n\n'\
            'Effective Date: ' + li['Effective Date']+'\n'\
            'Submitted Rate to CACI: $' + li['Submitted Rate to CACI']+'\n'\
            'SLA: ' + li['SLA']+'\n'\
            'Resource ID: ' + li['Resource ID']+'\n'\
            'CLIN: ' + li['CLIN']+'\n'\
            '---\n' \
            'CACI Internal/FYI Kimberly\n' \
            'Resource ID: ' + li['Resource ID']+'\n'\
            'Cost Center: ' + li['Cost Center']+'\n'\
            '--------------------------------------------------------------------\n\n' \
            'Regards,\nAaron Davis | AGDS Lead Staffing Coordinator\nITDAS.PMO@CACI.com\nIntel Applications Services\n1540 Conference Center Drive | Suite 100 | Chantilly, Va 20151\n' \
            'Office: 703.667.9197 | Cell: 202.329.3537\nAaron.Davis@CACI.com | ww.caci.com'

mail.Send()